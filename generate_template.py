from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

wb = Workbook()

# Sheet 0: Shift Definitions
ws_shifts = wb.active
ws_shifts.title = "Shift Definitions"

shift_headers = ["team", "code", "meaning", "start_time", "end_time", "timezone", "show_on_roster_page", "remark"]
ws_shifts.append(shift_headers)

shift_data = [
    ["L1", "A", "00:00-07:00", "00:00", "07:00", "HKT", "Y", ""],
    ["L1", "B", "06:30-15:30", "06:30", "15:30", "HKT", "Y", ""],
    ["L1", "D", "15:30-00:30", "15:30", "00:30", "HKT", "Y", ""],
    ["AP L2", "DS", "Day Shift", "09:30", "18:30", "HKT", "Y", ""],
    ["AP L2", "NS", "Night Shift", "18:30", "09:30", "HKT", "Y", ""],
    ["AP L2", "AP-M", "AP Morning Shift", "06:00", "15:00", "HKT", "Y", ""],
    ["AP L2", "AP-E", "AP Evening Shift", "15:00", "00:00", "HKT", "Y", ""],
    ["AP L2", "AP-D", "AP Day Shift", "09:00", "18:00", "HKT", "Y", ""],
]

for row in shift_data:
    ws_shifts.append(row)

# Style headers
header_font = Font(bold=True)
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws_shifts[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for row in ws_shifts.iter_rows(min_row=2, max_row=ws_shifts.max_row, min_col=1, max_col=len(shift_headers)):
    for cell in row:
        cell.border = thin_border

# Adjust column widths
for col in ws_shifts.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws_shifts.column_dimensions[column].width = max_length + 2

# Sheet 1: Staff Shifts
ws_staff = wb.create_sheet("Staff Shifts")

staff_headers = ["name", "staff_id", "team", "region", "contact", "notes"]
for day in range(1, 32):
    staff_headers.append(str(day))

ws_staff.append(staff_headers)

staff_data = [
    ["John Doe", "1001", "L1", "China", "", "", "A", "A", "B", "B", "D", "D", "A", "A", "B", "B", "D", "D", "A", "A", "B", "B", "D", "D", "A", "A", "B", "B", "D", "D", "A", "A", "B", "B", "D", "D"],
    ["Jane Smith", "1002", "AP L2", "China", "", "", "DS", "DS", "NS", "NS", "DS", "DS", "NS", "NS", "DS", "DS", "NS", "NS", "DS", "DS", "NS", "NS", "DS", "DS", "NS", "NS", "DS", "DS", "NS", "NS", "DS", "DS", "NS", "NS", "DS", "DS"],
    ["Mike Johnson", "1003", "AP L2", "China", "", "", "AP-M", "AP-M", "AP-E", "AP-E", "AP-D", "AP-D", "AP-M", "AP-M", "AP-E", "AP-E", "AP-D", "AP-D", "AP-M", "AP-M", "AP-E", "AP-E", "AP-D", "AP-D", "AP-M", "AP-M", "AP-E", "AP-E", "AP-D", "AP-D", "AP-M", "AP-M", "AP-E", "AP-E", "AP-D", "AP-D"],
]

for row in staff_data:
    ws_staff.append(row)

for cell in ws_staff[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for row in ws_staff.iter_rows(min_row=2, max_row=ws_staff.max_row, min_col=1, max_col=len(staff_headers)):
    for cell in row:
        cell.border = thin_border

for col in ws_staff.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws_staff.column_dimensions[column].width = max(max_length + 2, 5)

# Sheet 2: Color Definitions
ws_colors = wb.create_sheet("Color Definitions")

color_headers = ["code", "color_name", "rgb", "hex"]
ws_colors.append(color_headers)

color_data = [
    ["A", "Orange", "255 165 0", "#FFA500"],
    ["B", "DarkOrange", "255 140 0", "#FF8C00"],
    ["D", "OrangeRed", "255 69 0", "#FF4500"],
    ["DS", "RoyalBlue", "65 105 225", "#4169E1"],
    ["NS", "MidnightBlue", "25 25 112", "#191970"],
    ["AP-M", "LightSkyBlue", "135 206 250", "#87CEFA"],
    ["AP-E", "SteelBlue", "70 130 180", "#4682B4"],
    ["AP-D", "CornflowerBlue", "100 149 237", "#6495ED"],
]

for row in color_data:
    ws_colors.append(row)

for cell in ws_colors[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center')

for row in ws_colors.iter_rows(min_row=2, max_row=ws_colors.max_row, min_col=1, max_col=len(color_headers)):
    for cell in row:
        cell.border = thin_border

for col in ws_colors.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws_colors.column_dimensions[column].width = max_length + 2

# Save the workbook
output_path = "/Users/lzn/Documents/trae_projects/support/support-roster-server/src/main/resources/roster.xlsx"
wb.save(output_path)
print(f"Template saved to: {output_path}")
print(f"Sheets: {wb.sheetnames}")
